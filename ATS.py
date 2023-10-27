#!/usr/bin/env python

import datetime
import logging
import sys
import openpyxl
from openpyxl.utils import get_column_letter
import calendar
import exrex
import re
import pyinputplus as pyip
from typing import Pattern

''' Logging Config.'''

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

formatter = logging.Formatter('%(asctime)s: %(name)s - [%(levelname)s] - %(message)s')

file_handler = logging.FileHandler('ATS.log', mode='w')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(formatter)

# stream_handler = logging.StreamHandler()
# stream_handler.setFormatter(formatter)
# stream_handler.setLevel(logging.INFO)

logger.addHandler(file_handler)
# logger.addHandler(stream_handler)


'''Backend Functions'''


def find_sheet(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]

    else:
        print('Sheet not found.')
        logger.exception("Sheet not found: ")


def find_cell(cell_content, sheet):
    for col in sheet.iter_cols():
        for cell in col:
            if isinstance(cell_content, int):
                if cell.value == cell_content:
                    return cell
            if isinstance(cell_content, str):
                if isinstance(cell.value, str):
                    if cell.value.lower() == cell_content.lower():
                        return cell
            if isinstance(cell_content, Pattern):
                if cell_content.search(cell.value):
                    return cell
    logger.debug('Cell not found.')


def find_start_date(classes_start, month_start):
    if classes_start > month_start:
        return classes_start
    else:
        return month_start


def find_end_date(classes_stop, month_end):
    if classes_stop < month_end:
        return classes_stop
    else:
        return month_end


def find_working_dates(first_date, last_date, workdays):

    day_delta = datetime.timedelta(days=1)
    date_range = (last_date - first_date).days
    dates = []

    for i in range(date_range + 1):
        date = first_date + i * day_delta
        if date.strftime('%A') in workdays:
            dates.append(date)

    return dates


def find_working_hours(start, end):

    class_time = end - start
    class_time_hours = class_time.seconds / 3600

    return class_time_hours


def convert_time_to_dt(hour_minutes):
    # could account for different time formats

    time_formats = ['%H:%M', '%I:%M%p', '%H.%M', '%I.%M%p']

    for i in time_formats:
        try:
            dt_obj = datetime.datetime.strptime(hour_minutes, i)

            return dt_obj

        except ValueError:
            continue


def convert_date_to_dt(date):
    date_formats = list(exrex.generate(r'^(%d)(/(%m|%b|%B)/|-(%m|%b|%B)-|\.(%m|%b|%B)\.)(%y|%Y)$'))

    for i in date_formats:
        try:
            dt_obj = datetime.datetime.strptime(date, i)

            return dt_obj

        except ValueError:
            continue


def cell_filler(sheet, first_row, cell_count, column, content):

    if isinstance(content, list):
        for i in range(cell_count):
            sheet.cell(row=first_row + i, column=column).value = content[i]

    else:
        for i in range(cell_count):
            sheet.cell(row=first_row + i, column=column).value = content


def find_available_row(sheet, target_row, target_column):

    while sheet.cell(row=target_row, column=target_column).value:
        target_row += 1
    return target_row


def find_cell_content_in_column(sheet, target_row, target_column):

    cell_content = []

    while sheet.cell(row=target_row, column=target_column).value:
        cell_content.append(sheet.cell(row=target_row, column=target_column).value)
        target_row += 1
    return cell_content


def transpose(list_matrix):
    transposed_list = []

    for i in range(len(list_matrix[0])):
        row = []
        for item in list_matrix:
            row.append(item[i])
        transposed_list.append(row)

    return transposed_list


def remove_row(indexes, content_by_row):

    if len(indexes) == 1:
        del content_by_row[indexes[0]]
        return content_by_row

    if len(indexes) > 1:
        content_to_remove = [content_by_row[i] for i in indexes]
        new_list = [i for i in content_by_row if i not in content_to_remove]
        return new_list


def append_file_name(addition, file):

    file_type_re = re.compile(r'(\.[a-z]{2,4})$')
    file_type_suffix = file_type_re.search(file)
    file_name_updated = re.sub(file_type_re, addition + file_type_suffix.group(), file)
    return file_name_updated


def scale_cell_to_content(sheet, target_cell):

    content = target_cell.value
    target_col_letter = get_column_letter(target_cell.column)

    num_of_chars = float(len(content))
    if num_of_chars > 8.43:
        sheet.column_dimensions[target_col_letter].width = num_of_chars
    else:
        sheet.column_dimensions[target_col_letter].width = 8.43


'''CLI FUNCTIONS'''


def ask_sheet(workbook):
    sheet_choice = pyip.inputMenu(wb.sheetnames,
                                  prompt='Multiple sheets detected in excel document.\n'
                                         'Please choose the sheet you want to work with:\n',
                                  numbered=True)
    return sheet_choice


def ask_add_or_remove():
    class_choice = pyip.inputMenu(['Add a new class', 'Remove a class'],
                                  numbered=True)
    return class_choice


def ask_once_or_repeating():
    repeating_choice = pyip.inputMenu(['Once off', 'Repeating'],
                                      prompt='What type of class would you like to add? \n',
                                      numbered=True)
    return repeating_choice


def ask_class_name():
    name = pyip.inputStr(prompt='Please enter class name: ')
    return name


def ask_one_off_date():

    date_formats = list(exrex.generate(r'^(%d)(/(%m|%b|%B)/|-(%m|%b|%B)-|\.(%m|%b|%B)\.)(%y|%Y)$'))

    date = pyip.inputDate(prompt='Please enter class date: ', formats=date_formats)

    return [date]


def ask_from_to_dates():

    date_formats = list(exrex.generate(r'^(%d)(/(%m|%b|%B)/|-(%m|%b|%B)-|\.(%m|%b|%B)\.)(%y|%Y)$'))

    from_date = pyip.inputDate(prompt='Please enter class starting date: ', formats=date_formats)
    to_date = pyip.inputDate(prompt='Please enter class ending date: ', formats=date_formats)

    return [from_date, to_date]


def ask_class_time():

    st_morning = datetime.time(hour=8, minute=30)
    et_morning = datetime.time(hour=11)

    st_afternoon = datetime.time(hour=13, minute=30)
    et_afternoon = datetime.time(hour=16)

    class_times = []

    class_time_choice = pyip.inputMenu(
        ['Usual morning slot', 'Usual afternoon slot', 'Custom time'],
        prompt='Please choose class time:\n',
        numbered=True)

    if class_time_choice == 'Usual morning slot':
        class_times = st_morning, et_morning

    if class_time_choice == 'Usual afternoon slot':
        class_times = st_afternoon, et_afternoon

    if class_time_choice == 'Custom time':
        time_formats = ['%H:%M', '%I:%M%p', '%H.%M', '%I.%M%p']
        st_custom = pyip.inputTime(prompt='Please input start time: ', formats=time_formats)
        et_custom = pyip.inputTime(prompt='Please input end time: ', formats=time_formats)
        class_times = st_custom, et_custom

    class_times = [datetime.datetime.combine(datetime.date.min, i) for i in class_times]

    return class_times


def ask_for_working_days():

    day_names_full = list(calendar.day_name)
    day_names_abrvs = list(calendar.day_abbr)
    day_variations = dict(zip(day_names_full, day_names_abrvs))

    workdays = pyip.inputStr(prompt="Please enter the name of the days you'll be working: ")
    if ',' in workdays:
        workdays = workdays.split(',')
        workdays = [i.strip() for i in workdays]
    else:
        workdays = workdays.split(' ')

    workdays = [i for i in workdays if i != '']

    workdays_formatted = []

    for i in workdays:
        if i.title() in day_variations.keys():
            workdays_formatted.append(i.title())
        if i.title() in day_variations.values():
            [workdays_formatted.append(k) for k, v in day_variations.items() if v == i.title()]

    return workdays_formatted


def ask_row_to_remove(content_by_row):

    [print(i + 1, row) for i, row in enumerate(content_by_row)]

    choice_to_remove = input('Enter the number(s) of the classes you want to remove: ')

    if ',' in choice_to_remove:
        choice_to_remove = choice_to_remove.split(',')
        choice_to_remove = [i.strip() for i in choice_to_remove]
    else:
        choice_to_remove = choice_to_remove.split(' ')

    choice_to_remove = [int(i) for i in choice_to_remove if i != '']
    remove_index = [i - 1 for i in choice_to_remove]
    return remove_index


'''PROGRAM'''

logger.debug('Program Start')

target_excel = sys.argv[1]
wb = openpyxl.load_workbook(target_excel, data_only=True)
sh = wb.active



print('Welcome to AutoTimeSheet')

while True:

    if len(wb.sheetnames) > 1:
        sh_choice = ask_sheet(wb)
        sh = find_sheet(wb, sh_choice)

    date_re = re.compile(r'\d{2}/\d{2}/\d{4}')
    date_cell = find_cell(date_re, sh)
    from_to_dates = date_re.findall(date_cell.value)
    start_of_month = convert_date_to_dt(from_to_dates[0]).date()
    end_of_month = convert_date_to_dt(from_to_dates[1]).date()

    header_cells = {
                    'Class': find_cell('Class', sh),
                    'Date': find_cell('Date', sh),
                    'Start time': find_cell('Start time', sh),
                    'End time': find_cell('End time', sh),
                    'Hour': find_cell('Hour', sh)
                    }

    header_columns = {k: v.column for k, v in header_cells.items()}
    header_row = find_cell('Class', sh).row
    first_entry_row = header_row + 1

    cell_content_by_column = [
        find_cell_content_in_column(sh, first_entry_row, header_columns[k])
        for k in header_cells.keys()]
    logger.debug(cell_content_by_column)

    cell_content_by_column[1] = [
        convert_date_to_dt(i).date()
        if isinstance(i, str)
        else i.date()
        for i in cell_content_by_column[1]]

    cell_content_by_column[2] = [
        convert_time_to_dt(i)
        if isinstance(i, str)
        else datetime.datetime.combine(datetime.date.min, i)
        for i in cell_content_by_column[2]]

    cell_content_by_column[3] = [
        convert_time_to_dt(i)
        if isinstance(i, str)
        else datetime.datetime.combine(datetime.date.min, i)
        for i in cell_content_by_column[3]]

    cell_content_by_row = transpose(cell_content_by_column)

    add_or_remove = ask_add_or_remove()

    if add_or_remove == 'Add a new class':
        one_or_more = ask_once_or_repeating()

        if one_or_more == 'Once off':
            class_name = ask_class_name()
            class_dates = ask_one_off_date()
            classes_start_time, classes_end_time = ask_class_time()

            if class_dates[0] > start_of_month:
                if class_dates[0] < end_of_month:
                    working_dates = [class_dates[0]]
                date_count = 1
                logger.debug('WORKING DATES ONE OFF')
                logger.debug(working_dates)

        if one_or_more == 'Repeating':
            class_name = ask_class_name()
            class_dates = ask_from_to_dates()
            classes_start_time, classes_end_time = ask_class_time()
            working_days = ask_for_working_days()\

            first_class_of_month = find_start_date(class_dates[0], start_of_month)
            last_class_of_month = find_end_date(class_dates[1], end_of_month)
            working_dates = find_working_dates(first_class_of_month, last_class_of_month, working_days)
            logger.debug('WORKING DATES REPEATING')
            for i in working_dates:
                logger.debug(i)
            date_count = len(working_dates)

        working_hours = find_working_hours(classes_start_time, classes_end_time)

        for i in range(date_count):
            cell_content_by_row.append(
                [class_name, working_dates[i], classes_start_time, classes_end_time, working_hours])

    if add_or_remove == 'Remove a class':
        index_to_remove = ask_row_to_remove(cell_content_by_row)
        cell_content_by_row = remove_row(index_to_remove, cell_content_by_row)

    print('Updating Excel...')

    logger.debug(cell_content_by_row)
    cell_content_by_row.sort(key=lambda x: x[1])
    rows_of_content = len(cell_content_by_row)

    cell_content_by_column = transpose(cell_content_by_row)

    cell_content_by_column[1] = [
        datetime.datetime.strftime(i, '%d/%m/%Y')
        for i in cell_content_by_column[1]]

    cell_content_by_column[2] = [
        datetime.datetime.strftime(i, '%H:%M')
        for i in cell_content_by_column[2]]

    cell_content_by_column[3] = [
        datetime.datetime.strftime(i, '%H:%M')
        for i in cell_content_by_column[3]]

    for i, k in enumerate(header_columns.keys()):
        cell_filler(sh, first_entry_row, rows_of_content, header_columns[k], cell_content_by_column[i])

    first_available_row = find_available_row(sh, first_entry_row, header_columns['Date'])
    rows_in_sheet = first_available_row - first_entry_row
    num_rows_to_clear = rows_in_sheet - rows_of_content
    clear_from = first_available_row - num_rows_to_clear

    for k in header_columns.keys():
        cell_filler(sh, clear_from, num_rows_to_clear, header_columns[k], None)

    further_additions = pyip.inputYesNo(prompt='Would you like to make additional changes to the excel? ')
    if further_additions == 'yes':
        continue
    if further_additions == 'no':
        break

first_hour_row = sh.cell(row=first_entry_row, column=header_columns['Hour'])
last_hour_row = sh.cell(row=first_available_row-1, column=header_columns['Hour'])
total_header = sh.cell(header_cells['Hour'].row, header_cells['Hour'].column + 2)
total_content = sh.cell(total_header.row, total_header.column + 1)
total_header.value = 'Total Hours:'
total_content.value = '=SUM(%s:%s)' % (first_hour_row.coordinate, last_hour_row.coordinate)
scale_cell_to_content(sh, total_header)

new_name = append_file_name('[COMPLETED]', target_excel)
wb.save(new_name)

print('Excel Updated.')

